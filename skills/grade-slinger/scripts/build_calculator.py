#!/usr/bin/env python3
"""
Build a bespoke grade calculator Excel file for a course.

This is called by the Grade Slinger skill after it has parsed the syllabus and
extracted the gradebook structure. The skill builds a `course_spec` dict and
passes it to `build_calculator`. The result is a styled .xlsx file with one
input cell per individual grade item, formulas that calculate effective course
contributions, a letter-grade lookup, and hover comments on key cells.

See references/excel-calculator-spec.md for the design rules this script enforces.

Usage from Python (the skill will typically import and call directly):

    from build_calculator import build_calculator

    course_spec = {
        "course_name": "MBA 561 Consumer Behavior",
        "instructor": "Dr. Stefanie Robinson",
        "aggregation": "natural",  # or "weighted"
        "total_points": 1000,       # required for natural; ignored for weighted
        "categories": [
            {
                "name": "Exams",
                "course_percent": 70.0,
                "items": [
                    {"name": "Exam 1", "max": 350},
                    {"name": "Final Exam", "max": 350},
                ],
            },
            ...
        ],
        "letter_grades": [
            {"letter": "A+", "min_percent": 97.0},
            {"letter": "A",  "min_percent": 93.0},
            ...
            {"letter": "F",  "min_percent": 0.0},
        ],
    }

    build_calculator(course_spec, "/path/to/MBA561-grade-calculator.xlsx")

CLI usage (for testing):

    python3 build_calculator.py spec.json output.xlsx

where spec.json is a file matching the structure above.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- Styling constants ----------

NCSU_RED = "CC0000"
INPUT_YELLOW = "FFF3CD"
SUBTOTAL_GRAY = "E8E8E8"
SECTION_HEADER_GRAY = "F8F8F8"
RESULT_GREEN = "D4EDDA"

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _bold(color: str = "000000", size: int = 11) -> Font:
    return Font(bold=True, color=color, size=size)


# ---------- Main builder ----------

def build_calculator(
    course_spec: dict[str, Any],
    output_path: str | Path,
    mode: str = "full",
) -> Path:
    """Build the grade calculator workbook. Returns the output Path.

    mode:
      - "full": one input row per individual grade item (default)
      - "abbreviated": one input row per category (instructor enters the category total),
        plus one row per standalone item (top-level items not in a category)
    """
    if mode not in ("full", "abbreviated"):
        raise ValueError(f"mode must be 'full' or 'abbreviated', got {mode!r}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Calculator"

    course_name = course_spec.get("course_name", "Course")
    instructor = course_spec.get("instructor", "")
    aggregation = course_spec.get("aggregation", "weighted").lower()
    total_points = float(course_spec.get("total_points") or 0)
    categories = course_spec.get("categories", [])
    standalone_items = course_spec.get("standalone_items", [])
    letter_grades = course_spec.get("letter_grades", _default_letter_grades())

    # ----- Title and header rows -----
    mode_label = "Full" if mode == "full" else "Abbreviated"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{course_name} — Grade Calculator ({mode_label})"
    ws["A1"].font = Font(bold=True, color=NCSU_RED, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    if instructor:
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Instructor: {instructor}"
        ws["A2"].font = Font(italic=True, color="666666", size=10)

    ws.merge_cells("A3:D3")
    if mode == "abbreviated":
        ws["A3"] = (
            "Abbreviated version: enter one total per category (rather than every individual item). "
            "Hover over cells with red corners to see notes and formulas."
        )
    else:
        ws["A3"] = "Hover over cells with red corners to see notes and formulas."
    ws["A3"].font = Font(italic=True, color="666666", size=10)

    # ----- Column headers (row 5) -----
    HEADER_ROW = 5
    headers = ["Category / Item", "Max Grade", "Student Score", "Course Points"]
    header_notes = [
        "Category names and individual grade items. Items are indented under their category.",
        "The maximum points possible for each grade item.",
        "Enter the student's score for each item here. Yellow cells = input cells.",
        "Each item's contribution to the final course percentage. Calculated automatically.",
    ]
    for col_idx, (h, note) in enumerate(zip(headers, header_notes), start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        cell.font = _bold(color="FFFFFF")
        cell.fill = _fill(NCSU_RED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        cell.comment = Comment(note, "Grade Slinger")
    ws.row_dimensions[HEADER_ROW].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    # ----- Category and item rows -----
    current_row = HEADER_ROW + 1
    contribution_rows: list[int] = []  # rows whose Column D feeds into the final total

    for cat in categories:
        cat_name = cat["name"]
        cat_percent = float(cat.get("course_percent", 0))
        items = cat.get("items", [])
        # If the spec provides an explicit category max, use it; otherwise sum items
        cat_total_max = float(cat.get("max")) if cat.get("max") is not None else (
            sum(float(i["max"]) for i in items) if items else 0
        )

        if mode == "full":
            # ----- FULL MODE: one row per individual item -----

            # Category header row
            ws.cell(row=current_row, column=1, value=cat_name)
            ws.cell(row=current_row, column=1).font = _bold()
            ws.cell(row=current_row, column=1).comment = Comment(
                f"Category: {cat_name}\nEffective course weight: {cat_percent}%\n"
                f"Total available points in this category: {cat_total_max:.2f}",
                "Grade Slinger",
            )
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).fill = _fill(SECTION_HEADER_GRAY)
                ws.cell(row=current_row, column=col).border = THIN_BORDER
            current_row += 1

            item_start_row = current_row
            for item in items:
                item_name = item["name"]
                item_max = float(item["max"])

                cell_a = ws.cell(row=current_row, column=1, value=f"    {item_name}")
                cell_a.border = THIN_BORDER

                # Weighted-mean gradebooks give each item its own weight, which is
                # independent of max points. Honour an explicit course_percent when the
                # spec supplies one; otherwise fall back to the points-proportional
                # split used by natural/points-based courses.
                explicit_pct = item.get("course_percent")
                if explicit_pct is not None:
                    item_effective_pct = float(explicit_pct)
                    weight_basis = "explicit item weight"
                elif cat_total_max > 0:
                    item_effective_pct = (item_max / cat_total_max) * cat_percent
                    weight_basis = (
                        f"{item_max:.0f}/{cat_total_max:.0f} of category × {cat_percent}%"
                    )
                else:
                    item_effective_pct = 0
                    weight_basis = "no weight"
                cell_a.comment = Comment(
                    f"{item_name}\nMax points: {item_max:.2f}\n"
                    f"Effective course weight: {item_effective_pct:.3f}% "
                    f"({weight_basis})",
                    "Grade Slinger",
                )

                cell_b = ws.cell(row=current_row, column=2, value=item_max)
                cell_b.alignment = Alignment(horizontal="right")
                cell_b.border = THIN_BORDER

                cell_c = ws.cell(row=current_row, column=3)
                cell_c.fill = _fill(INPUT_YELLOW)
                cell_c.alignment = Alignment(horizontal="right")
                cell_c.border = THIN_BORDER

                score_ref = f"C{current_row}"
                max_ref = f"B{current_row}"
                if explicit_pct is not None:
                    # (score / max) × this item's effective course percentage
                    formula = (
                        f'=IF({score_ref}="","",'
                        f'{score_ref}/{max_ref}*{item_effective_pct})'
                    )
                    formula_note = (
                        f"Formula: score / {item_max:.0f} × {item_effective_pct:.4f}%"
                    )
                elif cat_total_max > 0:
                    formula = (
                        f'=IF({score_ref}="","",'
                        f'{score_ref}/{cat_total_max}*{cat_percent})'
                    )
                    formula_note = f"Formula: score / {cat_total_max:.0f} × {cat_percent}%"
                else:
                    formula = f'=IF({score_ref}="","",{score_ref}/{max_ref}*0)'
                    formula_note = "No weight assigned"
                cell_d = ws.cell(row=current_row, column=4, value=formula)
                cell_d.number_format = "0.000"
                cell_d.alignment = Alignment(horizontal="right")
                cell_d.border = THIN_BORDER
                cell_d.comment = Comment(
                    f"{formula_note}\n"
                    f"Max possible contribution: {item_effective_pct:.3f}%",
                    "Grade Slinger",
                )
                current_row += 1

            # Category subtotal row
            if items:
                subtotal_label = f"  {cat_name} subtotal"
                ws.cell(row=current_row, column=1, value=subtotal_label).font = _bold()
                ws.cell(row=current_row, column=2, value=cat_total_max).alignment = (
                    Alignment(horizontal="right")
                )
                ws.cell(row=current_row, column=2).font = _bold()
                ws.cell(row=current_row, column=4, value=(
                    f"=SUM(D{item_start_row}:D{current_row - 1})"
                )).number_format = "0.000"
                ws.cell(row=current_row, column=4).font = _bold()
                ws.cell(row=current_row, column=4).comment = Comment(
                    f"Sum of all items in {cat_name}.\n"
                    f"Max possible contribution: {cat_percent}%",
                    "Grade Slinger",
                )
                for col in range(1, 5):
                    ws.cell(row=current_row, column=col).fill = _fill(SUBTOTAL_GRAY)
                    ws.cell(row=current_row, column=col).border = THIN_BORDER
                contribution_rows.append(current_row)
                current_row += 1

            current_row += 1  # spacer

        else:
            # ----- ABBREVIATED MODE: one row per category (input the total) -----

            cell_a = ws.cell(row=current_row, column=1, value=cat_name)
            cell_a.font = _bold()
            cell_a.border = THIN_BORDER
            cell_a.comment = Comment(
                f"Category: {cat_name}\n"
                f"Effective course weight: {cat_percent}%\n"
                f"Enter the total points earned across all items in this category.\n"
                f"Total available: {cat_total_max:.2f}",
                "Grade Slinger",
            )

            cell_b = ws.cell(row=current_row, column=2, value=cat_total_max)
            cell_b.alignment = Alignment(horizontal="right")
            cell_b.font = _bold()
            cell_b.border = THIN_BORDER

            cell_c = ws.cell(row=current_row, column=3)
            cell_c.fill = _fill(INPUT_YELLOW)
            cell_c.alignment = Alignment(horizontal="right")
            cell_c.border = THIN_BORDER
            cell_c.comment = Comment(
                f"Enter the student's total points earned in {cat_name} "
                f"(out of {cat_total_max:.0f}).",
                "Grade Slinger",
            )

            score_ref = f"C{current_row}"
            if cat_total_max > 0:
                formula = (
                    f'=IF({score_ref}="","",'
                    f'{score_ref}/{cat_total_max}*{cat_percent})'
                )
            else:
                formula = f'=IF({score_ref}="","",0)'
            cell_d = ws.cell(row=current_row, column=4, value=formula)
            cell_d.number_format = "0.000"
            cell_d.alignment = Alignment(horizontal="right")
            cell_d.font = _bold()
            cell_d.border = THIN_BORDER
            cell_d.comment = Comment(
                f"Formula: category_total / {cat_total_max:.0f} × {cat_percent}%\n"
                f"Max possible contribution: {cat_percent}%",
                "Grade Slinger",
            )
            # Fill the category row with subtotal gray so it visually reads as a roll-up
            for col in range(1, 5):
                if col != 3:  # keep the input cell yellow
                    ws.cell(row=current_row, column=col).fill = _fill(SUBTOTAL_GRAY)

            contribution_rows.append(current_row)
            current_row += 1

    # ----- Standalone items (top-level items not in any category) -----
    if standalone_items:
        # Section header
        ws.cell(row=current_row, column=1, value="Standalone Items").font = _bold()
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).fill = _fill(SECTION_HEADER_GRAY)
            ws.cell(row=current_row, column=col).border = THIN_BORDER
        current_row += 1

        for item in standalone_items:
            item_name = item["name"]
            item_max = float(item["max"])
            item_pct = float(item.get("course_percent", 0))

            cell_a = ws.cell(row=current_row, column=1, value=f"    {item_name}")
            cell_a.border = THIN_BORDER
            cell_a.comment = Comment(
                f"{item_name} (standalone, not in a category)\n"
                f"Max points: {item_max:.2f}\n"
                f"Effective course weight: {item_pct}%",
                "Grade Slinger",
            )

            cell_b = ws.cell(row=current_row, column=2, value=item_max)
            cell_b.alignment = Alignment(horizontal="right")
            cell_b.border = THIN_BORDER

            cell_c = ws.cell(row=current_row, column=3)
            cell_c.fill = _fill(INPUT_YELLOW)
            cell_c.alignment = Alignment(horizontal="right")
            cell_c.border = THIN_BORDER

            score_ref = f"C{current_row}"
            if item_max > 0:
                formula = (
                    f'=IF({score_ref}="","",{score_ref}/{item_max}*{item_pct})'
                )
            else:
                formula = f'=IF({score_ref}="","",0)'
            cell_d = ws.cell(row=current_row, column=4, value=formula)
            cell_d.number_format = "0.000"
            cell_d.alignment = Alignment(horizontal="right")
            cell_d.border = THIN_BORDER
            cell_d.comment = Comment(
                f"Formula: score / {item_max:.0f} × {item_pct}%\n"
                f"Max possible contribution: {item_pct}%",
                "Grade Slinger",
            )

            contribution_rows.append(current_row)
            current_row += 1

        current_row += 1  # spacer

    # ----- Final percentage and letter grade -----
    final_row = current_row + 1
    ws.cell(row=final_row, column=1, value="Final Course Percentage").font = _bold()
    if contribution_rows:
        subtotal_refs = ",".join(f"D{r}" for r in contribution_rows)
        ws.cell(row=final_row, column=4, value=f"=SUM({subtotal_refs})")
    else:
        ws.cell(row=final_row, column=4, value=0)
    ws.cell(row=final_row, column=4).number_format = "0.00"
    ws.cell(row=final_row, column=4).font = _bold(size=12)
    ws.cell(row=final_row, column=4).fill = _fill(RESULT_GREEN)
    ws.cell(row=final_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=final_row, column=1).fill = _fill(RESULT_GREEN)

    letter_row = final_row + 1
    ws.cell(row=letter_row, column=1, value="Final Letter Grade").font = _bold()
    # Build a lookup table on a hidden sheet to support VLOOKUP
    lookup_sheet_name = "_LetterGrades"
    lookup_ws = wb.create_sheet(lookup_sheet_name)
    lookup_ws.sheet_state = "hidden"
    # VLOOKUP needs ascending lookup values, so build it that way
    sorted_letters = sorted(letter_grades, key=lambda x: x["min_percent"])
    for idx, lg in enumerate(sorted_letters, start=1):
        lookup_ws.cell(row=idx, column=1, value=lg["min_percent"])
        lookup_ws.cell(row=idx, column=2, value=lg["letter"])
    lookup_range = (
        f"{lookup_sheet_name}!A1:B{len(sorted_letters)}"
    )
    ws.cell(row=letter_row, column=4, value=(
        f'=IFERROR(VLOOKUP(D{final_row},{lookup_range},2,TRUE),"")'
    ))
    ws.cell(row=letter_row, column=4).font = _bold(size=12, color=NCSU_RED)
    ws.cell(row=letter_row, column=4).fill = _fill(RESULT_GREEN)
    ws.cell(row=letter_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=letter_row, column=1).fill = _fill(RESULT_GREEN)
    # Tooltip with the full grade scale
    scale_text_lines = ["Grade scale:"]
    for lg in sorted(letter_grades, key=lambda x: -x["min_percent"]):
        scale_text_lines.append(f"  {lg['letter']}: {lg['min_percent']}%+")
    ws.cell(row=letter_row, column=4).comment = Comment(
        "\n".join(scale_text_lines), "Grade Slinger"
    )

    # ----- Instruction footer -----
    footer_row = letter_row + 3
    ws.cell(row=footer_row, column=1,
            value="Instructions:").font = _bold()
    if mode == "abbreviated":
        instructions = [
            "1. Enter the student's TOTAL points earned for each category (yellow cells, Column C).",
            "   For example, if a student earned 165 of 175 in Participation, enter 165.",
            "2. The Course Points column auto-calculates each category's contribution.",
            "3. The Final Course Percentage and Letter Grade update automatically.",
            "4. Leave a cell blank to exclude that category from the calculation.",
            "5. Hover over any cell with a red corner to see notes and formulas.",
        ]
    else:
        instructions = [
            "1. Enter the student's score for each item in the yellow cells (Column C).",
            "2. The Course Points column auto-calculates each item's contribution.",
            "3. The Final Course Percentage and Letter Grade update automatically.",
            "4. Leave a cell blank to exclude that item from the calculation.",
            "5. Hover over any cell with a red corner to see notes and formulas.",
        ]
    for i, line in enumerate(instructions, start=1):
        ws.cell(row=footer_row + i, column=1, value=line).font = Font(
            italic=True, color="666666", size=10
        )

    # Freeze panes below header row
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _default_letter_grades() -> list[dict[str, Any]]:
    """Default Moodle plus/minus scheme."""
    return [
        {"letter": "A+", "min_percent": 97.0},
        {"letter": "A",  "min_percent": 93.0},
        {"letter": "A-", "min_percent": 90.0},
        {"letter": "B+", "min_percent": 87.0},
        {"letter": "B",  "min_percent": 83.0},
        {"letter": "B-", "min_percent": 80.0},
        {"letter": "C+", "min_percent": 77.0},
        {"letter": "C",  "min_percent": 73.0},
        {"letter": "C-", "min_percent": 70.0},
        {"letter": "D+", "min_percent": 67.0},
        {"letter": "D",  "min_percent": 63.0},
        {"letter": "D-", "min_percent": 60.0},
        {"letter": "F",  "min_percent": 0.0},
    ]


# ---------- CLI for testing ----------

def main() -> int:
    if len(sys.argv) not in (3, 4):
        print(
            "Usage: python3 build_calculator.py <spec.json> <output.xlsx> [full|abbreviated]",
            file=sys.stderr,
        )
        return 1
    spec_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    mode = sys.argv[3] if len(sys.argv) == 4 else "full"
    spec = json.loads(spec_path.read_text())
    out = build_calculator(spec, output_path, mode=mode)
    print(f"Wrote {out} (mode={mode})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
